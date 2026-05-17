"""
handlers/admin_input.py
=======================
Admin message handler — full state machine for all multi-step flows:
channel adding, link generation, clone setup, category settings,
upload manager, auto-forward, manga tracker, broadcast, user mgmt, etc.
"""
import asyncio
import csv
import html
import json
from datetime import datetime
from io import StringIO, BytesIO
from typing import Optional

from telegram import Update, InlineKeyboardButton, InlineKeyboardMarkup
from telegram.constants import ParseMode
from telegram.ext import ContextTypes

from core.config import ADMIN_ID, OWNER_ID, BOT_USERNAME, DEFAULT_CAPTION
from core.logging_setup import logger
from core.helpers import safe_send_message, safe_edit_text, UserFriendlyError
from core.buttons import _btn, _back_btn, _close_btn, bold_button, _back_kb
from core.text_utils import b, bq, code, e, small_caps
from core.state_machine import (
    user_states, upload_progress,
    ADD_CHANNEL_USERNAME, ADD_CHANNEL_TITLE, PENDING_CHANNEL_POST,
    GENERATE_LINK_IDENTIFIER, GENERATE_LINK_TITLE,
    ADD_CLONE_TOKEN, SET_BACKUP_CHANNEL,
    PENDING_BROADCAST, PENDING_BROADCAST_OPTIONS, PENDING_BROADCAST_CONFIRM,
    SCHEDULE_BROADCAST_DATETIME, SCHEDULE_BROADCAST_MSG,
    SET_CATEGORY_CAPTION, SET_CATEGORY_BRANDING, SET_CATEGORY_BUTTONS,
    SET_CATEGORY_THUMBNAIL, SET_WATERMARK_TEXT,
    UPLOAD_SET_CAPTION, UPLOAD_SET_SEASON, UPLOAD_SET_EPISODE,
    UPLOAD_SET_TOTAL, UPLOAD_SET_CHANNEL,
    AF_ADD_CONNECTION_SOURCE, AF_ADD_CONNECTION_TARGET,
    AU_ADD_MANGA_TITLE, AU_ADD_MANGA_TARGET, AU_CUSTOM_INTERVAL,
    BAN_USER_INPUT, UNBAN_USER_INPUT, DELETE_USER_INPUT, SEARCH_USER_INPUT,
    CW_SET_TEXT, CW_SET_BUTTONS, BroadcastMode,
)


async def handle_admin_message(
    update: Update, context: ContextTypes.DEFAULT_TYPE
) -> None:
    """Handle all messages from admin in conversation states."""
    if not update.effective_user:
        return
    uid = update.effective_user.id
    if uid not in (ADMIN_ID, OWNER_ID):
        return
    if uid not in user_states:
        return
    if not update.message:
        return

    state = user_states[uid]
    text = update.message.text or ""
    chat_id = update.effective_chat.id

    # Forward-based states: route to photo handler which handles them
    _channel_states_that_accept_fwd = {
        PENDING_CHANNEL_POST, ADD_CHANNEL_USERNAME,
        AF_ADD_CONNECTION_SOURCE, AF_ADD_CONNECTION_TARGET, AU_ADD_MANGA_TARGET,
    }
    if state in _channel_states_that_accept_fwd:
        _fwd_src = (
            getattr(update.message, "forward_from_chat", None)
            or (getattr(update.message, "forward_origin", None)
                and getattr(update.message.forward_origin, "chat", None))
        )
        if _fwd_src:
            from handlers.admin_photo import handle_admin_photo
            await handle_admin_photo(update, context)
            return

    # Delete trigger and old prompt
    try:
        await update.message.delete()
    except Exception:
        pass
    msg_id = context.user_data.pop("bot_prompt_message_id", None)
    if msg_id:
        try:
            await context.bot.delete_message(chat_id, msg_id)
        except Exception:
            pass

    # Cancel
    if text.strip().lower() in ("/cancel", "cancel"):
        user_states.pop(uid, None)
        from handlers.admin_panel import send_admin_menu
        await send_admin_menu(chat_id, context)
        return

    # ── Channel states ─────────────────────────────────────────────────────────

    if state == ADD_CHANNEL_USERNAME:
        uname = text.strip()
        if not uname.startswith("@") and not uname.lstrip("-").isdigit():
            msg = await safe_send_message(
                context.bot, chat_id,
                b("❌ Invalid format. Use @username or numeric channel ID. Try again:"),
                reply_markup=InlineKeyboardMarkup([[bold_button("🔙 Cancel", callback_data="manage_force_sub")]]),
            )
            context.user_data["bot_prompt_message_id"] = msg.message_id if msg else None
            return
        lookup = int(uname) if uname.lstrip("-").isdigit() else (uname if uname.startswith("@") else f"@{uname}")
        try:
            tg_chat = await context.bot.get_chat(lookup)
            is_private = not tg_chat.username
            # Store @username for public channels so the join URL works;
            # numeric ID for private channels (invite link generated later).
            context.user_data["new_ch_uname"] = (
                f"@{tg_chat.username}" if tg_chat.username else str(tg_chat.id)
            )
            context.user_data["new_ch_title"]      = tg_chat.title
            context.user_data["new_ch_is_private"] = is_private
            context.user_data["new_ch_jbr"]        = False  # default; admin can pick later
            context.user_data["new_ch_id"]         = tg_chat.id  # numeric chat ID for JBR lookup
            user_states[uid] = ADD_CHANNEL_TITLE
            ch_info = f"<b>Channel:</b> {e(tg_chat.title)}\n<b>ID:</b> <code>{tg_chat.id}</code>"
            if tg_chat.username:
                ch_info += f"\n<b>Username:</b> @{e(tg_chat.username)}"
            else:
                ch_info += "\n<b>Type:</b> Private channel"
            private_note = (
                "\n\n" + b("Is this a Join-Request (JBR) channel?") +
                "\n<i>If users must request to join, use the JBR button below.</i>"
                if is_private else ""
            )
            jbr_row = [
                bold_button("🔔 JBR Mode", callback_data="new_ch_jbr_yes"),
                bold_button("📢 Direct Join", callback_data="new_ch_jbr_no"),
            ] if is_private else []
            btn_rows = ([jbr_row] if jbr_row else []) + [[bold_button("🔙 Cancel", callback_data="manage_force_sub")]]
            msg = await safe_send_message(
                context.bot, chat_id,
                b("✅ Channel found!") + "\n\n" + bq(ch_info) + private_note + "\n\n"
                + b("Send a display title, or /skip to use the channel name:"),
                reply_markup=InlineKeyboardMarkup(btn_rows),
            )
            context.user_data["bot_prompt_message_id"] = msg.message_id if msg else None
        except Exception as exc:
            msg = await safe_send_message(
                context.bot, chat_id,
                b("❌ Cannot access that channel.\n\n") + bq(
                    b("Make sure:\n• Bot is admin in the channel\n• Username/ID is correct\n\nError: ")
                ) + code(e(str(exc)[:120])),
                reply_markup=InlineKeyboardMarkup([[bold_button("🔙 Cancel", callback_data="manage_force_sub")]]),
            )
            context.user_data["bot_prompt_message_id"] = msg.message_id if msg else None
        return

    if state == ADD_CHANNEL_TITLE:
        uname = context.user_data.get("new_ch_uname")
        if not uname:
            user_states.pop(uid, None)
            await safe_send_message(context.bot, chat_id, b("Session expired. Start over."))
            return
        title = text.strip()
        if title.lower() == "/skip":
            title = context.user_data.get("new_ch_title", uname)

        # Detect if this is a private channel (numeric ID, no public @username).
        # Also read the stored join_by_request flag if set by ADD_CHANNEL_USERNAME.
        join_by_request = bool(context.user_data.pop("new_ch_jbr", False))
        stored_invite   = context.user_data.pop("new_ch_invite_link", None) or ""
        channel_id_int  = context.user_data.pop("new_ch_id", None)
        is_private      = uname.lstrip("-").isdigit()

        # For private channels with no stored invite link, generate one now.
        if is_private and not stored_invite:
            try:
                chat_obj = await context.bot.get_chat(int(uname))
                if join_by_request:
                    lnk_obj = await context.bot.create_chat_invite_link(
                        int(uname), creates_join_request=True
                    )
                else:
                    lnk_obj = await context.bot.export_chat_invite_link(int(uname))
                stored_invite = getattr(lnk_obj, "invite_link", lnk_obj) or ""
            except Exception as _ex:
                logger.debug(f"Could not auto-generate invite link for {uname}: {_ex}")

        from database_dual import add_force_sub_channel
        add_force_sub_channel(
            uname, title,
            join_by_request=join_by_request,
            invite_link=stored_invite or None,
            channel_id=channel_id_int,
        )
        jbr_note = " (🔔 Join-Request mode)" if join_by_request else ""
        link_note = f"\n<i>Invite link stored.</i>" if stored_invite else (
            "\n⚠️ <i>No invite link — bot may not be admin there. Use /setinvite.</i>"
            if is_private else ""
        )
        await safe_send_message(
            context.bot, chat_id,
            b(f"✅ Added {e(title)} ({e(uname)}) as force-sub channel!") + jbr_note + link_note,
        )
        user_states.pop(uid, None)
        from handlers.admin_panel import send_admin_menu
        await send_admin_menu(chat_id, context)
        return

    # ── Link generation ────────────────────────────────────────────────────────

    if state == GENERATE_LINK_IDENTIFIER:
        identifier = text.strip()
        lookup = int(identifier) if identifier.lstrip("-").isdigit() else (identifier if identifier.startswith("@") else f"@{identifier}")
        try:
            tg_chat = await context.bot.get_chat(lookup)
            context.user_data["gen_ch_id"] = tg_chat.id
            context.user_data["gen_ch_title"] = tg_chat.title or str(tg_chat.id)
            user_states[uid] = GENERATE_LINK_TITLE
            msg = await safe_send_message(
                context.bot, chat_id,
                b(small_caps(f"📢 channel: {tg_chat.title or str(tg_chat.id)}")) + "\n\n"
                + bq(
                    b(small_caps("send the filter keyword / title for this link:\n\n"))
                    + small_caps("this becomes both the link title and the filter trigger.\n")
                    + small_caps("send /skip to use the channel name: ")
                    + f"<code>{e(tg_chat.title or '')}</code>"
                ),
                reply_markup=InlineKeyboardMarkup([[bold_button("🔙 Cancel", callback_data="admin_back")]]),
            )
            context.user_data["bot_prompt_message_id"] = msg.message_id if msg else None
        except Exception as exc:
            msg = await safe_send_message(
                context.bot, chat_id,
                b("❌ Cannot access that channel.\n\n") + bq(
                    b("Error: ") + code(e(str(exc)[:100]))
                ),
                reply_markup=InlineKeyboardMarkup([[bold_button("🔙 Cancel", callback_data="admin_back")]]),
            )
            context.user_data["bot_prompt_message_id"] = msg.message_id if msg else None
        return

    if state == GENERATE_LINK_TITLE:
        title = text.strip()
        if title.lower() == "/skip":
            title = context.user_data.get("gen_ch_title", "")
        ch_id = context.user_data.get("gen_ch_id")
        if not ch_id:
            user_states.pop(uid, None)
            await safe_send_message(context.bot, chat_id, b(small_caps("session expired. start over.")))
            return
        try:
            from database_dual import generate_link_id
            link_id = generate_link_id(
                channel_username=ch_id, user_id=uid,
                never_expires=False, channel_title=title,
                source_bot_username=BOT_USERNAME,
            )
            deep_link = f"https://t.me/{BOT_USERNAME}?start={link_id}"
            await safe_send_message(
                context.bot, chat_id,
                b(small_caps("✅ link generated!")) + "\n\n"
                + bq(code(deep_link)) + "\n\n"
                + bq(
                    b(small_caps("🎌 filter auto-active!")) + "\n"
                    + small_caps("keyword: ") + f"<code>{e(title)}</code>\n"
                    + small_caps("when users type this title in any group, "
                                 "they get the poster + join button automatically.")
                ),
                reply_markup=_back_kb(),
            )
        except Exception as exc:
            await safe_send_message(
                context.bot, chat_id,
                b(small_caps("❌ error generating link: ")) + code(e(str(exc)[:200])),
            )
        user_states.pop(uid, None)
        context.user_data.pop("gen_ch_id", None)
        context.user_data.pop("gen_ch_title", None)
        return

    # ── Clone token ────────────────────────────────────────────────────────────

    if state == ADD_CLONE_TOKEN:
        token = text.strip()
        from handlers.clones import _register_clone_token
        await _register_clone_token(update, context, token)
        user_states.pop(uid, None)
        return

    # ── Send poster to channel (custom channel input) ─────────────────────────

    if isinstance(state, str) and state.startswith("AWAITING_SEND_TO_CHANNEL:"):
        parts    = state.split(":", 2)
        src_chat_s = parts[1] if len(parts) > 1 else ""
        src_msg_s  = parts[2] if len(parts) > 2 else ""
        dest_input = text.strip()
        try:
            src_chat_i = int(src_chat_s)
            src_msg_i  = int(src_msg_s)
            try:
                dest = int(dest_input)
            except ValueError:
                dest = dest_input  # @username
            import json as _json
            from database_dual import get_setting
            raw  = get_setting(f"last_poster_{uid}", "")
            pdat = _json.loads(raw) if raw else {}
            cap  = pdat.get("caption", "")
            await context.bot.copy_message(
                chat_id=dest,
                from_chat_id=src_chat_i,
                message_id=src_msg_i,
                caption=cap,
                parse_mode="HTML",
            )
            await safe_send_message(context.bot, chat_id, b("✅ Poster sent to channel!"))
        except Exception as exc:
            from core.helpers import UserFriendlyError
            await safe_send_message(context.bot, chat_id,
                b("❌ Failed: ") + code(e(str(exc)[:120])))
        user_states.pop(uid, None)
        return


    if isinstance(state, str) and (state.startswith("chatbot_key:") or state.startswith("chatbot_new_set:")):
        from handlers.chatbot_panel import handle_chatbot_key_input
        handled = await handle_chatbot_key_input(update, context, state)
        if handled:
            user_states.pop(uid, None)
        return

    # ── Backup channel ─────────────────────────────────────────────────────────

    if state == SET_BACKUP_CHANNEL:
        from database_dual import set_setting
        set_setting("backup_channel_url", text.strip())
        await safe_send_message(context.bot, chat_id, b(f"✅ Backup channel URL set: {e(text.strip())}"))
        user_states.pop(uid, None)
        from handlers.admin_panel import send_admin_menu
        await send_admin_menu(chat_id, context)
        return

    # ── Broadcast ──────────────────────────────────────────────────────────────

    if state == PENDING_BROADCAST:
        context.user_data["broadcast_message"] = (update.message.chat_id, update.message.message_id)
        user_states[uid] = PENDING_BROADCAST_OPTIONS
        keyboard = [
            [bold_button("Normal", callback_data="broadcast_mode_normal"),
             bold_button("Silent", callback_data="broadcast_mode_silent")],
            [bold_button("Auto-Delete 24h", callback_data="broadcast_mode_auto_delete"),
             bold_button("Pin", callback_data="broadcast_mode_pin")],
            [bold_button("Schedule", callback_data="broadcast_schedule"),
             bold_button("🔙 Cancel", callback_data="admin_back")],
        ]
        msg = await safe_send_message(
            context.bot, chat_id, b("✅ Message received! Choose broadcast mode:"),
            reply_markup=InlineKeyboardMarkup(keyboard),
        )
        context.user_data["bot_prompt_message_id"] = msg.message_id if msg else None
        return

    if state == PENDING_BROADCAST_CONFIRM and text.strip().lower() in ("/confirm", "confirm"):
        msg_data = context.user_data.get("broadcast_message")
        mode = context.user_data.get("broadcast_mode", BroadcastMode.NORMAL)
        if not msg_data:
            await safe_send_message(context.bot, chat_id, b("❌ Broadcast message lost. Start over."))
            user_states.pop(uid, None)
            return
        user_states.pop(uid, None)
        msg_chat_id, msg_id = msg_data
        from handlers.admin_panel import _do_broadcast
        asyncio.create_task(_do_broadcast(context, chat_id, msg_chat_id, msg_id, mode))
        return

    # ── Category settings ──────────────────────────────────────────────────────

    category = context.user_data.get("editing_category", "")

    if state == SET_CATEGORY_CAPTION:
        from handlers.post_gen import update_category_field
        update_category_field(category, "caption_template", text.strip())
        await safe_send_message(context.bot, chat_id, b(f"✅ Caption template for {e(category)} updated!"))
        user_states.pop(uid, None)
        from handlers.admin_panel import send_admin_menu
        await send_admin_menu(chat_id, context)
        return

    if state == SET_CATEGORY_BRANDING:
        from handlers.post_gen import update_category_field, show_category_settings_menu
        val = text.strip()
        if update_category_field(category, "branding", val):
            await safe_send_message(
                context.bot, chat_id,
                b(f"✅ Branding for {e(category.upper())} updated!") + "\n"
                + bq(code(e(val[:120])) if val else b("(cleared)")),
            )
        else:
            await safe_send_message(context.bot, chat_id, b("❌ Failed to save branding."))
        user_states.pop(uid, None)
        await show_category_settings_menu(context, chat_id, category, None)
        return

    if state == SET_CATEGORY_BUTTONS:
        from handlers.post_gen import update_category_field, show_category_settings_menu
        lines = text.strip().split("\n")
        buttons_list = []
        for line in lines:
            if " - " in line:
                parts = line.split(" - ", 1)
                buttons_list.append({"text": parts[0].strip(), "url": parts[1].strip()})
        if update_category_field(category, "buttons", json.dumps(buttons_list)):
            await safe_send_message(context.bot, chat_id, b(f"✅ {len(buttons_list)} button(s) configured!"))
        else:
            await safe_send_message(context.bot, chat_id, b("❌ Failed to save buttons."))
        user_states.pop(uid, None)
        await show_category_settings_menu(context, chat_id, category, None)
        return

    if state == SET_CATEGORY_THUMBNAIL:
        from handlers.post_gen import update_category_field, show_category_settings_menu
        val = "" if text.strip().lower() in ("default", "none", "remove", "clear") else text.strip()
        update_category_field(category, "thumbnail_url", val)
        await safe_send_message(context.bot, chat_id, b(f"✅ Thumbnail {'reset' if not val else 'updated'}!"))
        user_states.pop(uid, None)
        await show_category_settings_menu(context, chat_id, category, None)
        return

    if state == SET_WATERMARK_TEXT:
        from handlers.post_gen import update_category_field, show_category_settings_menu
        update_category_field(category, "watermark_text", text.strip())
        await safe_send_message(context.bot, chat_id, b("✅ Watermark text set!"))
        user_states.pop(uid, None)
        await show_category_settings_menu(context, chat_id, category, None)
        return

    # ── Upload ─────────────────────────────────────────────────────────────────

    if state == UPLOAD_SET_CAPTION:
        from handlers.upload import save_upload_progress, show_upload_menu
        upload_field = context.user_data.pop("upload_field", None)
        if upload_field == "anime_name":
            upload_progress["anime_name"] = text.strip()
        else:
            upload_progress["base_caption"] = text
        await save_upload_progress()
        await safe_send_message(context.bot, chat_id, b("✅ Updated!"))
        user_states.pop(uid, None)
        await show_upload_menu(chat_id, context)
        return

    if state == UPLOAD_SET_SEASON:
        from handlers.upload import save_upload_progress, show_upload_menu
        try:
            upload_progress["season"] = int(text.strip())
            upload_progress["video_count"] = 0
            await save_upload_progress()
            await safe_send_message(context.bot, chat_id, b(f"✅ Season set to {upload_progress['season']}"))
        except ValueError:
            await safe_send_message(context.bot, chat_id, b("❌ Invalid number."))
            return
        user_states.pop(uid, None)
        await show_upload_menu(chat_id, context)
        return

    if state == UPLOAD_SET_EPISODE:
        from handlers.upload import save_upload_progress, show_upload_menu
        try:
            upload_progress["episode"] = int(text.strip())
            upload_progress["video_count"] = 0
            await save_upload_progress()
            await safe_send_message(context.bot, chat_id, b(f"✅ Episode set to {upload_progress['episode']}"))
        except ValueError:
            await safe_send_message(context.bot, chat_id, b("❌ Invalid number."))
            return
        user_states.pop(uid, None)
        await show_upload_menu(chat_id, context)
        return

    if state == UPLOAD_SET_TOTAL:
        from handlers.upload import save_upload_progress, show_upload_menu
        try:
            upload_progress["total_episode"] = int(text.strip())
            await save_upload_progress()
            await safe_send_message(context.bot, chat_id, b(f"✅ Total set to {upload_progress['total_episode']}"))
        except ValueError:
            await safe_send_message(context.bot, chat_id, b("❌ Invalid number."))
            return
        user_states.pop(uid, None)
        await show_upload_menu(chat_id, context)
        return

    if state == UPLOAD_SET_CHANNEL:
        from handlers.upload import save_upload_progress, show_upload_menu
        identifier = text.strip()
        try:
            tg_chat = await context.bot.get_chat(identifier)
            upload_progress["target_chat_id"] = tg_chat.id
            await save_upload_progress()
            await safe_send_message(context.bot, chat_id, b(f"✅ Target set to: {e(tg_chat.title)}"))
        except Exception as exc:
            await safe_send_message(context.bot, chat_id, UserFriendlyError.get_user_message(exc))
            return
        user_states.pop(uid, None)
        await show_upload_menu(chat_id, context)
        return

    # ── Auto-forward ───────────────────────────────────────────────────────────

    if state == AF_ADD_CONNECTION_SOURCE:
        identifier = text.strip()
        lookup = int(identifier) if identifier.lstrip("-").isdigit() else (identifier if identifier.startswith("@") else f"@{identifier}")
        try:
            tg_chat = await context.bot.get_chat(lookup)
            context.user_data["af_source_id"] = tg_chat.id
            context.user_data["af_source_uname"] = tg_chat.username
            user_states[uid] = AF_ADD_CONNECTION_TARGET
            msg = await safe_send_message(
                context.bot, chat_id,
                b(f"✅ Source: {e(tg_chat.title)}") + "\n\n"
                + bq(b("Step 2/2: Send the TARGET channel @username or ID:")),
                reply_markup=InlineKeyboardMarkup([[bold_button("🔙 Cancel", callback_data="admin_autoforward")]]),
            )
            context.user_data["bot_prompt_message_id"] = msg.message_id if msg else None
        except Exception as exc:
            await safe_send_message(context.bot, chat_id, b(f"❌ Cannot access: {e(str(exc)[:100])}"))
        return

    if state == AF_ADD_CONNECTION_TARGET:
        identifier = text.strip()
        lookup = int(identifier) if identifier.lstrip("-").isdigit() else (identifier if identifier.startswith("@") else f"@{identifier}")
        try:
            tg_chat = await context.bot.get_chat(lookup)
            src_id = context.user_data.get("af_source_id")
            src_uname = context.user_data.get("af_source_uname", "")
            if not src_id:
                await safe_send_message(context.bot, chat_id, b("Session expired."))
                user_states.pop(uid, None)
                return
            from database_dual import db_manager
            with db_manager.get_cursor() as cur:
                cur.execute("""
                    INSERT INTO auto_forward_connections
                        (source_chat_id, source_chat_username, target_chat_id,
                         target_chat_username, active)
                    VALUES (%s, %s, %s, %s, TRUE)
                    ON CONFLICT DO NOTHING
                """, (src_id, src_uname, tg_chat.id, tg_chat.username))
            await safe_send_message(
                context.bot, chat_id,
                b("✅ Auto-forward connection created!") + "\n\n"
                + bq(
                    b("Source: ") + code(str(src_id)) + "\n"
                    + b("Target: ") + code(str(tg_chat.id)) + " — " + e(tg_chat.title)
                ),
            )
        except Exception as exc:
            await safe_send_message(context.bot, chat_id, b(f"❌ Error: {e(str(exc)[:100])}"))
        user_states.pop(uid, None)
        from handlers.admin_panel import send_admin_menu
        await send_admin_menu(chat_id, context)
        return

    # ── Manga tracker ──────────────────────────────────────────────────────────

    if state == AU_ADD_MANGA_TITLE:
        from api.mangadex import MangaDexClient
        from api.anilist import AniListClient
        title = text.strip()
        results = MangaDexClient.search_manga(title, limit=5)
        if not results:
            al = AniListClient.search_manga(title)
            if al:
                al_t = (al.get("title") or {})
                al_title_str = al_t.get("romaji") or al_t.get("english") or title
                results = MangaDexClient.search_manga(al_title_str, limit=5)
        if not results:
            await safe_send_message(context.bot, chat_id, b("❌ No manga found on MangaDex."))
            return
        keyboard = []
        for manga in results[:5]:
            attrs = manga.get("attributes", {}) or {}
            titles = attrs.get("title", {}) or {}
            manga_title = titles.get("en") or next(iter(titles.values()), "Unknown")
            keyboard.append([bold_button(manga_title[:40], callback_data=f"mdex_track_{manga['id']}")])
        keyboard.append([bold_button("🔙 Cancel", callback_data="admin_autoupdate")])
        await safe_send_message(
            context.bot, chat_id, b("📚 Select the manga to track:"),
            reply_markup=InlineKeyboardMarkup(keyboard),
        )
        user_states.pop(uid, None)
        return

    if state == AU_CUSTOM_INTERVAL:
        try:
            mins = int(text.strip())
            if mins < 1:
                raise ValueError
        except ValueError:
            await safe_send_message(
                context.bot, chat_id, b("❌ Send a valid number of minutes (e.g. 15):"),
                reply_markup=InlineKeyboardMarkup([[bold_button("🔙 Cancel", callback_data="admin_autoupdate")]]),
            )
            return
        context.user_data["au_manga_interval"] = mins
        user_states[uid] = AU_ADD_MANGA_TARGET
        t = context.user_data.get("au_manga_title", "Unknown")
        m = context.user_data.get("au_manga_mode", "latest")
        await safe_send_message(
            context.bot, chat_id,
            b(f"📚 {e(t)}") + f"\n<b>Mode:</b> {m} | <b>Interval:</b> {mins} min\n\n"
            + bq(b("Send the target channel @username, numeric ID, or forward a post:")),
            reply_markup=InlineKeyboardMarkup([[bold_button("🔙 Cancel", callback_data="admin_autoupdate")]]),
        )
        return

    if state == AU_ADD_MANGA_TARGET:
        identifier = text.strip()
        manga_id = context.user_data.get("au_manga_id")
        manga_title = context.user_data.get("au_manga_title", "Unknown")
        manga_mode = context.user_data.get("au_manga_mode", "latest")
        manga_interval = context.user_data.get("au_manga_interval", 60)
        if not manga_id:
            await safe_send_message(context.bot, chat_id, b("Session expired. Please start over."))
            user_states.pop(uid, None)
            return
        try:
            _ident = identifier.strip()
            if _ident.lstrip("-").isdigit():
                _ident = int(_ident)
            elif not _ident.startswith("@"):
                _ident = f"@{_ident}"
            tg_chat = await context.bot.get_chat(_ident)
            from api.mangadex import MangaTracker, MangaDexClient
            success = MangaTracker.add_tracking(manga_id, manga_title, tg_chat.id)
            if success:
                await safe_send_message(
                    context.bot, chat_id,
                    b(f"✅ Now tracking: {e(manga_title)}") + "\n\n"
                    + bq(
                        f"<b>Channel:</b> {e(tg_chat.title or str(tg_chat.id))}\n"
                        f"<b>Mode:</b> {manga_mode.title()}\n"
                        f"<b>Interval:</b> {manga_interval} min\n\n"
                        + b("New chapters will be sent automatically.")
                    ),
                )
            else:
                await safe_send_message(context.bot, chat_id, b("❌ Failed to add tracking. Make sure bot is admin."))
        except Exception as exc:
            await safe_send_message(context.bot, chat_id, b(f"❌ Error: {e(str(exc)[:100])}"))
            return
        user_states.pop(uid, None)
        for k in ("au_manga_id", "au_manga_title", "au_manga_mode", "au_manga_interval"):
            context.user_data.pop(k, None)
        from handlers.admin_panel import send_admin_menu
        await send_admin_menu(chat_id, context)
        return

    # ── Channel welcome states ─────────────────────────────────────────────────

    if state == CW_SET_TEXT:
        ch_id = context.user_data.get("cw_editing_channel")
        if not ch_id:
            user_states.pop(uid, None)
            await safe_send_message(context.bot, chat_id, b("session expired."))
            return
        from database_dual import set_channel_welcome
        set_channel_welcome(ch_id, welcome_text=text.strip())
        user_states.pop(uid, None)
        await safe_send_message(
            context.bot, chat_id, b(small_caps("✅ welcome text saved!")),
            reply_markup=InlineKeyboardMarkup([[
                InlineKeyboardButton(small_caps("✏️ edit more"), callback_data=f"cw_edit_{ch_id}"),
                _back_btn("admin_channel_welcome"),
            ]]),
        )
        return

    if state == CW_SET_BUTTONS:
        ch_id = context.user_data.get("cw_editing_channel")
        if not ch_id:
            user_states.pop(uid, None)
            await safe_send_message(context.bot, chat_id, b("session expired."))
            return
        lines = text.strip().split("\n")
        btns = []
        for line in lines:
            if " - " in line:
                parts = line.split(" - ", 1)
                btns.append({"text": parts[0].strip(), "url": parts[1].strip()})
        from database_dual import set_channel_welcome
        set_channel_welcome(ch_id, buttons=btns)
        user_states.pop(uid, None)
        await safe_send_message(
            context.bot, chat_id, b(small_caps(f"✅ {len(btns)} button(s) saved!")),
            reply_markup=InlineKeyboardMarkup([[
                InlineKeyboardButton(small_caps("✏️ edit more"), callback_data=f"cw_edit_{ch_id}"),
                _back_btn("admin_channel_welcome"),
            ]]),
        )
        return

    if isinstance(state, str) and state == "CW_WAITING_CHANNEL_ID":
        identifier = text.strip()
        lookup = int(identifier) if identifier.lstrip("-").isdigit() else (identifier if identifier.startswith("@") else f"@{identifier}")
        try:
            tg_chat = await context.bot.get_chat(lookup)
            from database_dual import set_channel_welcome
            set_channel_welcome(tg_chat.id, enabled=True, welcome_text="", added_by=uid)
            context.user_data["cw_editing_channel"] = tg_chat.id
            user_states.pop(uid, None)
            await safe_send_message(
                context.bot, chat_id,
                b(small_caps(f"✅ channel registered: {tg_chat.title or str(tg_chat.id)}")),
                reply_markup=InlineKeyboardMarkup([[
                    InlineKeyboardButton(small_caps("⚙️ configure"), callback_data=f"cw_edit_{tg_chat.id}"),
                    _back_btn("admin_channel_welcome"),
                ]]),
            )
        except Exception as exc:
            await safe_send_message(
                context.bot, chat_id,
                b(small_caps("❌ cannot access that channel.")) + "\n"
                + bq(code(e(str(exc)[:100]))),
                reply_markup=InlineKeyboardMarkup([[_back_btn("admin_channel_welcome")]]),
            )
        return

    if isinstance(state, str) and state == "CW_AWAITING_IMAGE":
        ch_id = context.user_data.get("cw_editing_channel")
        if ch_id and text.strip().startswith("http"):
            from database_dual import set_channel_welcome
            set_channel_welcome(ch_id, image_url=text.strip(), image_file_id="")
            user_states.pop(uid, None)
            await safe_send_message(context.bot, chat_id, b(small_caps("✅ welcome image url saved!")))
        else:
            await safe_send_message(context.bot, chat_id, b(small_caps("send a photo or image url (https://).")))
        return

    # ── User management ────────────────────────────────────────────────────────

    if state == BAN_USER_INPUT:
        from database_dual import resolve_target_user_id, ban_user
        target = resolve_target_user_id(text.strip())
        if target and target not in (ADMIN_ID, OWNER_ID):
            ban_user(target)
            await safe_send_message(context.bot, chat_id, b(f"🚫 User {code(str(target))} banned."))
        else:
            await safe_send_message(context.bot, chat_id, b("❌ Cannot ban admin or user not found."))
        user_states.pop(uid, None)
        return

    if state == UNBAN_USER_INPUT:
        from database_dual import resolve_target_user_id, unban_user
        target = resolve_target_user_id(text.strip())
        if target:
            unban_user(target)
            await safe_send_message(context.bot, chat_id, b(f"✅ User {code(str(target))} unbanned."))
        else:
            await safe_send_message(context.bot, chat_id, b("❌ User not found."))
        user_states.pop(uid, None)
        return

    if state == DELETE_USER_INPUT:
        try:
            from database_dual import db_manager
            target_uid = int(text.strip())
            if target_uid in (ADMIN_ID, OWNER_ID):
                await safe_send_message(context.bot, chat_id, b("⚠️ Cannot delete admin/owner."))
            else:
                with db_manager.get_cursor() as cur:
                    cur.execute("DELETE FROM users WHERE user_id = %s", (target_uid,))
                await safe_send_message(context.bot, chat_id, b(f"✅ User {code(str(target_uid))} deleted."))
        except Exception as exc:
            await safe_send_message(context.bot, chat_id, b(f"❌ Error: {code(e(str(exc)[:100]))}"))
        user_states.pop(uid, None)
        return

    if state == SEARCH_USER_INPUT:
        from database_dual import resolve_target_user_id, get_user_info_by_id
        target = resolve_target_user_id(text.strip())
        if target:
            user_info = get_user_info_by_id(target)
            if user_info:
                u_id, u_uname, u_fname, u_lname, u_joined, u_banned = user_info
                name = f"{u_fname or ''} {u_lname or ''}".strip() or "N/A"
                await safe_send_message(
                    context.bot, chat_id,
                    b("👤 User Found:") + "\n\n"
                    + bq(
                        f"<b>ID:</b> {code(str(u_id))}\n"
                        f"<b>Name:</b> {e(name)}\n"
                        f"<b>Username:</b> {'@' + e(u_uname) if u_uname else '—'}\n"
                        f"<b>Joined:</b> {str(u_joined)[:16]}\n"
                        f"<b>Status:</b> {'🚫 Banned' if u_banned else '✅ Active'}"
                    ),
                )
            else:
                await safe_send_message(context.bot, chat_id, b(f"❌ No user found with ID {target}."))
        else:
            await safe_send_message(context.bot, chat_id, b("❌ User not found."))
        user_states.pop(uid, None)
        return

    # ── Scheduled broadcast ────────────────────────────────────────────────────

    if state == SCHEDULE_BROADCAST_DATETIME:
        try:
            dt = datetime.strptime(text.strip(), "%Y-%m-%d %H:%M")
            context.user_data["schedule_dt"] = dt
            from core.state_machine import SCHEDULE_BROADCAST_MSG
            user_states[uid] = SCHEDULE_BROADCAST_MSG
            msg = await safe_send_message(
                context.bot, chat_id,
                b(f"📅 Scheduled for: {dt.strftime('%d %b %Y %H:%M')} UTC") + "\n\n"
                + bq(b("Now send the message to broadcast:")),
            )
            context.user_data["bot_prompt_message_id"] = msg.message_id if msg else None
        except ValueError:
            await safe_send_message(context.bot, chat_id, b("❌ Invalid format. Use: YYYY-MM-DD HH:MM"))
        return

    if state == SCHEDULE_BROADCAST_MSG:
        dt = context.user_data.get("schedule_dt")
        if not dt:
            user_states.pop(uid, None)
            await safe_send_message(context.bot, chat_id, b("❌ Session expired. Start over."))
            return
        try:
            from database_dual import db_manager
            with db_manager.get_cursor() as cur:
                cur.execute("""
                    INSERT INTO scheduled_broadcasts (admin_id, message_text, execute_at, status)
                    VALUES (%s, %s, %s, 'pending')
                """, (uid, text.strip(), dt))
        except Exception as exc:
            await safe_send_message(context.bot, chat_id, b("❌ Error scheduling: ") + code(e(str(exc)[:200])))
            user_states.pop(uid, None)
            return
        await safe_send_message(
            context.bot, chat_id,
            b(f"✅ Broadcast scheduled for {dt.strftime('%d %b %Y %H:%M')} UTC!"),
            reply_markup=_back_kb(),
        )
        user_states.pop(uid, None)
        return

    # ── ENV variable editing ───────────────────────────────────────────────────

    if isinstance(state, str) and state.startswith("AWAITING_ENV_"):
        env_key = state[len("AWAITING_ENV_"):]
        user_states.pop(uid, None)
        from database_dual import set_setting
        if text.strip().lower() == "reset":
            try:
                from database_dual import _pg_run
                _pg_run("DELETE FROM bot_settings WHERE key = %s", (f"env_{env_key}",))
            except Exception:
                pass
            await safe_send_message(
                context.bot, chat_id, b(f"♻️ {e(env_key)} reset to .env default."),
                reply_markup=InlineKeyboardMarkup([[_back_btn("admin_env_panel"), _close_btn()]]),
            )
        else:
            val = text.strip()
            set_setting(f"env_{env_key}", val)
            if env_key == "TRANSITION_STICKER" and val:
                set_setting("loading_sticker_id", val)
                set_setting("loading_anim_enabled", "true")
            elif env_key == "TRANSITION_STICKER":
                set_setting("loading_sticker_id", "")
            await safe_send_message(
                context.bot, chat_id,
                b(f"✔️ {e(env_key)} updated.") + f"\n{code(e(val[:80]))}",
                reply_markup=InlineKeyboardMarkup([[_back_btn("admin_env_panel"), _close_btn()]]),
            )
        return

    # ── Misc string states ─────────────────────────────────────────────────────

    if isinstance(state, str) and state.startswith("AWAITING_WATERMARK_"):
        cat = state[len("AWAITING_WATERMARK_"):].lower()
        user_states.pop(uid, None)
        from handlers.post_gen import update_category_field, show_category_settings_menu
        if text.strip().lower() == "none":
            update_category_field(cat, "watermark_text", None)
            await safe_send_message(
                context.bot, chat_id, b(f"✔️ Watermark removed for {cat}."),
                reply_markup=InlineKeyboardMarkup([[_back_btn(f"cat_settings_{cat}"), _close_btn()]]),
            )
        else:
            parts = text.strip().split("|", 1)
            wm_text = parts[0].strip()
            wm_pos = parts[1].strip() if len(parts) > 1 else "center"
            update_category_field(cat, "watermark_text", wm_text)
            update_category_field(cat, "watermark_position", wm_pos)
            await safe_send_message(
                context.bot, chat_id,
                b(f"✔️ Watermark set: <i>{e(wm_text)}</i> @ {wm_pos}"),
                reply_markup=InlineKeyboardMarkup([[_back_btn(f"cat_settings_{cat}"), _close_btn()]]),
            )
        return

    for _filter_state, _col in (("AWAITING_USER_SEARCH", None),):
        if state == "AWAITING_USER_SEARCH":
            user_states.pop(uid, None)
            from database_dual import resolve_target_user_id, get_user_info_by_id
            target = resolve_target_user_id(text.strip())
            if target:
                info = get_user_info_by_id(target)
                if info:
                    uid2, uname, fname, lname, joined, banned = info
                    st = "🔴 BANNED" if banned else "🟢 Active"
                    await safe_send_message(
                        context.bot, chat_id,
                        b(f"USER INFO: {e(str(uid2))}") + "\n\n"
                        + bq(
                            f"<b>Name:</b> {e((fname or '') + ' ' + (lname or ''))}\n"
                            f"<b>Username:</b> @{e(uname or 'N/A')}\n"
                            f"<b>Joined:</b> {str(joined)[:10]}\n"
                            f"<b>Status:</b> {st}\n"
                            f"<b>ID:</b> <code>{uid2}</code>"
                        ),
                    )
                else:
                    await safe_send_message(context.bot, chat_id, b(f"❗ User {target} not found."))
            else:
                await safe_send_message(context.bot, chat_id, b("❗ Invalid user ID or username."))
            return

    if isinstance(state, str) and state in ("AWAITING_BAN_USER", "AWAITING_UNBAN_USER", "AWAITING_DELETE_USER"):
        user_states.pop(uid, None)
        from database_dual import resolve_target_user_id, ban_user, unban_user, db_manager
        if state == "AWAITING_BAN_USER":
            target = resolve_target_user_id(text.strip())
            if target and target not in (ADMIN_ID, OWNER_ID):
                ban_user(target)
                await safe_send_message(context.bot, chat_id, b(f"🔴 User <code>{target}</code> banned."),
                    reply_markup=InlineKeyboardMarkup([[_back_btn("user_management"), _close_btn()]]))
            else:
                await safe_send_message(context.bot, chat_id, b("❗ Cannot ban admin or user not found."))
        elif state == "AWAITING_UNBAN_USER":
            target = resolve_target_user_id(text.strip())
            if target:
                unban_user(target)
                await safe_send_message(context.bot, chat_id, b(f"🟢 User <code>{target}</code> unbanned."),
                    reply_markup=InlineKeyboardMarkup([[_back_btn("user_management"), _close_btn()]]))
            else:
                await safe_send_message(context.bot, chat_id, b("❗ User not found."))
        else:
            try:
                target = int(text.strip())
                with db_manager.get_cursor() as cur:
                    cur.execute("DELETE FROM users WHERE user_id = %s", (target,))
                await safe_send_message(context.bot, chat_id, b(f"✔️ User <code>{target}</code> deleted."),
                    reply_markup=InlineKeyboardMarkup([[_back_btn("user_management"), _close_btn()]]))
            except Exception as exc:
                await safe_send_message(context.bot, chat_id, b(f"❗ Error: {e(str(exc)[:100])}"))
        return

    if isinstance(state, str) and state == "AWAITING_LINK_EXPIRY":
        user_states.pop(uid, None)
        from database_dual import set_setting
        try:
            mins = int(text.strip())
            if 1 <= mins <= 1440:
                set_setting("link_expiry_override", str(mins))
                await safe_send_message(context.bot, chat_id, b(f"✔️ Link expiry set to {mins} minutes."),
                    reply_markup=InlineKeyboardMarkup([[_back_btn("admin_settings"), _close_btn()]]))
            else:
                await safe_send_message(context.bot, chat_id, b("❗ Must be between 1 and 1440 minutes."))
        except ValueError:
            await safe_send_message(context.bot, chat_id, b("❗ Send a valid number."))
        return

    # ── Panel image URL input ──────────────────────────────────────────────────

    if isinstance(state, str) and state == "AWAITING_PANEL_IMG_URLS":
        user_states.pop(uid, None)
        from core.panel_image import get_panel_db_images, save_panel_db_images
        from core.config import PANEL_DB_CHANNEL
        raw_entries = [v.strip() for v in text.replace(",", "\n").splitlines() if v.strip()]
        if not raw_entries:
            await safe_send_message(context.bot, chat_id, "❌ Nothing found.")
            return
        items = get_panel_db_images()
        added = 0
        errors = 0
        for entry in raw_entries:
            if entry.startswith("http"):
                items.append({"index": len(items) + 1, "msg_id": 0, "file_id": entry})
                added += 1
            elif PANEL_DB_CHANNEL:
                try:
                    sent = await context.bot.send_photo(
                        chat_id=PANEL_DB_CHANNEL, photo=entry,
                        caption=f"Panel image #{len(items)+1}",
                    )
                    fid = sent.photo[-1].file_id
                    items.append({"index": len(items) + 1, "msg_id": sent.message_id, "file_id": fid})
                    added += 1
                except Exception:
                    errors += 1
            else:
                items.append({"index": len(items) + 1, "msg_id": 0, "file_id": entry})
                added += 1
        if added:
            save_panel_db_images(items)
        err_note = f"\n⚠️ {errors} failed." if errors else ""
        await safe_send_message(
            context.bot, chat_id,
            b(f"✅ Added {added} panel image(s). Total: {len(items)}.") + err_note,
            reply_markup=InlineKeyboardMarkup([
                [InlineKeyboardButton("🖼 View", callback_data="panel_img_manage")],
                [_back_btn("admin_settings"), _close_btn()],
            ])
        )
        return

    if isinstance(state, str) and state == "AWAITING_FWD_CHAT":
        user_states.pop(uid, None)
        from database_dual import set_setting
        set_setting("fwd_source_chat", text.strip())
        await safe_send_message(context.bot, chat_id, b(f"✅ Forward source chat set: {e(text.strip())}"),
            reply_markup=InlineKeyboardMarkup([[_back_btn("fsub_fwd_source"), _close_btn()]]))
        return

    # ── Set main channel for "send to main channel" feature ───────────────────
    if isinstance(state, str) and state == "AWAITING_MAIN_CHANNEL_ID":
        user_states.pop(uid, None)
        raw = text.strip()
        # Accept forwarded message (channel post) or @username or numeric ID
        ch_id_str = raw
        if update.message and update.message.forward_from_chat:
            ch_id_str = str(update.message.forward_from_chat.id)
        elif raw.startswith("@"):
            try:
                ch = await context.bot.get_chat(raw)
                ch_id_str = str(ch.id)
            except Exception:
                await safe_send_message(context.bot, chat_id,
                    b(small_caps(f"❌ could not find channel: {e(raw)}")))
                return
        elif not raw.lstrip("-").isdigit():
            try:
                ch = await context.bot.get_chat(raw if raw.startswith("@") else f"@{raw}")
                ch_id_str = str(ch.id)
            except Exception:
                await safe_send_message(context.bot, chat_id,
                    b(small_caps(f"❌ invalid channel id: {e(raw)}")))
                return

        from database_dual import set_setting
        set_setting("main_channel_id", ch_id_str)
        await safe_send_message(
            context.bot, chat_id,
            b(small_caps("✅ main channel set!")) + "\n\n"
            + bq(small_caps(f"channel id: ") + f"<code>{e(ch_id_str)}</code>\n"
                 + small_caps("posters will be forwarded here when 'send to main channel' is tapped.")),
            reply_markup=InlineKeyboardMarkup([[_back_btn("admin_settings"), _close_btn()]]),
        )
        return

    if isinstance(state, str) and state == "AWAITING_FWD_MSGID":
        user_states.pop(uid, None)
        try:
            from database_dual import set_setting
            set_setting("fwd_source_msg_id", str(int(text.strip())))
            await safe_send_message(context.bot, chat_id, b(f"✅ Message ID set."),
                reply_markup=InlineKeyboardMarkup([[_back_btn("fsub_fwd_source"), _close_btn()]]))
        except ValueError:
            await safe_send_message(context.bot, chat_id, b("❌ Invalid message ID — must be a number."))
        return

    if isinstance(state, str) and state == "AWAITING_JOIN_BTN_TEXT":
        user_states.pop(uid, None)
        val = text.strip()
        if val:
            from database_dual import set_setting
            set_setting("env_JOIN_BTN_TEXT", val)
            await safe_send_message(
                context.bot, chat_id,
                b(small_caps("✅ join button text updated!")) + "\n" + bq(code(e(val))),
                reply_markup=InlineKeyboardMarkup([[_back_btn("admin_filter_poster"), _close_btn()]]),
            )
        else:
            await safe_send_message(context.bot, chat_id, b(small_caps("❌ empty text — not saved.")))
        return

    # ── DM auto-delete delay ──────────────────────────────────────────────────
    if isinstance(state, str) and state == "AWAITING_DM_DEL_DELAY":
        user_states.pop(uid, None)
        try:
            secs = int(text.strip())
            from database_dual import set_setting
            set_setting("auto_delete_dm_delay", str(max(0, secs)))
            await safe_send_message(
                context.bot, chat_id,
                b(small_caps(f"✅ dm auto-delete delay set: {secs}s" if secs else "✅ dm auto-delete disabled")),
                reply_markup=InlineKeyboardMarkup([[_back_btn("admin_settings"), _close_btn()]]),
            )
        except ValueError:
            await safe_send_message(context.bot, chat_id, b(small_caps("❌ invalid — send a number like 120")))
        return

    # ── GC auto-delete delay ──────────────────────────────────────────────────
    if isinstance(state, str) and state == "AWAITING_GC_DEL_DELAY":
        user_states.pop(uid, None)
        try:
            secs = int(text.strip())
            from database_dual import set_setting
            set_setting("auto_delete_gc_delay", str(max(0, secs)))
            await safe_send_message(
                context.bot, chat_id,
                b(small_caps(f"✅ gc auto-delete delay set: {secs}s" if secs else "✅ gc auto-delete disabled")),
                reply_markup=InlineKeyboardMarkup([[_back_btn("admin_settings"), _close_btn()]]),
            )
        except ValueError:
            await safe_send_message(context.bot, chat_id, b(small_caps("❌ invalid — send a number like 60")))
        return


    # ── Filter poster auto-delete (AWAITING_FILTER_AUTODEL) ───────────────────
    if isinstance(state, str) and state == "AWAITING_FILTER_AUTODEL":
        user_states.pop(uid, None)
        try:
            secs = int(text.strip())
            if secs < 0:
                raise ValueError
            from filter_poster import set_auto_delete_seconds, build_filter_poster_settings_keyboard, get_filter_poster_settings_text
            set_auto_delete_seconds(chat_id, secs)
            label = f"{secs}s ({secs // 60} min)" if secs else "disabled (never)"
            await safe_send_message(
                context.bot, chat_id,
                b(small_caps(f"✅ filter auto-delete set: {label}")),
                reply_markup=InlineKeyboardMarkup([[_back_btn("admin_filter_poster"), _close_btn()]]),
            )
        except ValueError:
            await safe_send_message(context.bot, chat_id,
                b(small_caps("❌ invalid — send seconds (e.g. 300 = 5 min, 0 = never)")))
        return

    # ── Filter poster link expiry (AWAITING_LINK_EXPIRY_FP) ──────────────────
    if isinstance(state, str) and state == "AWAITING_LINK_EXPIRY_FP":
        user_states.pop(uid, None)
        try:
            mins = int(text.strip())
            if mins < 0:
                raise ValueError
            from filter_poster import set_link_expiry_minutes
            set_link_expiry_minutes(chat_id, mins)
            label = f"{mins} min" if mins else "permanent (no expiry)"
            await safe_send_message(
                context.bot, chat_id,
                b(small_caps(f"✅ link expiry set: {label}")),
                reply_markup=InlineKeyboardMarkup([[_back_btn("admin_filter_poster"), _close_btn()]]),
            )
        except ValueError:
            await safe_send_message(context.bot, chat_id,
                b(small_caps("❌ invalid — send minutes (e.g. 5, 60, or 0 for permanent)")))
        return

    # ── Auto-forward destination (AWAITING_FWD_DEST) ─────────────────────────
    if isinstance(state, str) and state == "AWAITING_FWD_DEST":
        user_states.pop(uid, None)
        dest = text.strip()
        if dest:
            from database_dual import set_setting
            set_setting("autoforward_dest_chat", dest)
            await safe_send_message(
                context.bot, chat_id,
                b(small_caps(f"✅ auto-forward destination set: {e(dest)}")),
                reply_markup=InlineKeyboardMarkup([[_back_btn("fsub_fwd_source"), _close_btn()]]),
            )
        else:
            await safe_send_message(context.bot, chat_id, b(small_caps("❌ empty — send @username or chat ID")))
        return

    # ── Clone move links target (AWAITING_MOVE_LINKS) ────────────────────────
    if isinstance(state, str) and state == "AWAITING_MOVE_LINKS":
        user_states.pop(uid, None)
        target = text.strip()
        if not target:
            await safe_send_message(context.bot, chat_id, b(small_caps("❌ empty — send bot @username")))
            return
        try:
            from database_dual import get_all_links, update_link_bot
            links = get_all_links() or []
            moved = 0
            for link in links:
                try:
                    update_link_bot(link.get("link_id") or link.get("id"), target)
                    moved += 1
                except Exception:
                    pass
            await safe_send_message(
                context.bot, chat_id,
                b(small_caps(f"✅ moved {moved} links to {e(target)}")),
                reply_markup=InlineKeyboardMarkup([[_back_btn("manage_clones"), _close_btn()]]),
            )
        except Exception as exc_ml:
            await safe_send_message(context.bot, chat_id,
                b(small_caps(f"❌ error moving links: {e(str(exc_ml)[:80])}")))
        return

    # ── Import users file (AWAITING_IMPORT_USERS_FILE) ───────────────────────
    if isinstance(state, str) and state == "AWAITING_IMPORT_USERS_FILE":
        user_states.pop(uid, None)
        doc = message.document if message else None
        if not doc:
            await safe_send_message(context.bot, chat_id,
                b(small_caps("❌ please send a .csv or .xlsx file")))
            return
        try:
            file_obj = await context.bot.get_file(doc.file_id)
            import io, os
            buf = io.BytesIO()
            await file_obj.download_to_memory(buf)
            buf.seek(0)
            fname = doc.file_name or "import.csv"
            from database_dual import add_user
            imported = 0
            if fname.endswith(".csv"):
                import csv
                reader = csv.DictReader(io.StringIO(buf.read().decode("utf-8", errors="replace")))
                for row in reader:
                    uid_val = row.get("user_id") or row.get("id") or row.get("uid")
                    if uid_val:
                        try:
                            add_user(int(uid_val))
                            imported += 1
                        except Exception:
                            pass
            elif fname.endswith((".xlsx", ".xls")):
                import openpyxl
                wb = openpyxl.load_workbook(buf)
                ws = wb.active
                for row in ws.iter_rows(min_row=2, values_only=True):
                    uid_val = row[0]
                    if uid_val:
                        try:
                            add_user(int(uid_val))
                            imported += 1
                        except Exception:
                            pass
            await safe_send_message(
                context.bot, chat_id,
                b(small_caps(f"✅ imported {imported} users successfully")),
                reply_markup=InlineKeyboardMarkup([[_back_btn("user_management"), _close_btn()]]),
            )
        except Exception as exc_iuf:
            await safe_send_message(context.bot, chat_id,
                b(small_caps(f"❌ import failed: {e(str(exc_iuf)[:80])}")))
        return

    # ── Import links file (AWAITING_IMPORT_LINKS_FILE) ───────────────────────
    if isinstance(state, str) and state == "AWAITING_IMPORT_LINKS_FILE":
        user_states.pop(uid, None)
        doc = message.document if message else None
        if not doc:
            await safe_send_message(context.bot, chat_id,
                b(small_caps("❌ please send a .csv or .xlsx file")))
            return
        try:
            file_obj = await context.bot.get_file(doc.file_id)
            import io
            buf = io.BytesIO()
            await file_obj.download_to_memory(buf)
            buf.seek(0)
            fname = doc.file_name or "import.csv"
            from database_dual import save_link
            imported = 0
            rows = []
            if fname.endswith(".csv"):
                import csv
                reader = csv.DictReader(io.StringIO(buf.read().decode("utf-8", errors="replace")))
                rows = list(reader)
            elif fname.endswith((".xlsx", ".xls")):
                import openpyxl
                wb = openpyxl.load_workbook(buf)
                ws = wb.active
                headers = [c.value for c in next(ws.iter_rows(min_row=1, max_row=1))]
                for row in ws.iter_rows(min_row=2, values_only=True):
                    rows.append(dict(zip(headers, row)))
            for row in rows:
                try:
                    link_id   = str(row.get("link_id") or row.get("id") or "")
                    user_id   = int(row.get("user_id") or row.get("uid") or 0)
                    channel   = str(row.get("channel_id") or row.get("channel") or "")
                    link_url  = str(row.get("link") or row.get("url") or "")
                    if link_id and user_id:
                        save_link(link_id=link_id, user_id=user_id,
                                  channel_id=channel, link=link_url)
                        imported += 1
                except Exception:
                    pass
            await safe_send_message(
                context.bot, chat_id,
                b(small_caps(f"✅ imported {imported} links successfully")),
                reply_markup=InlineKeyboardMarkup([[_back_btn("admin_back"), _close_btn()]]),
            )
        except Exception as exc_ilf:
            await safe_send_message(context.bot, chat_id,
                b(small_caps(f"❌ import failed: {e(str(exc_ilf)[:80])}")))
        return

    # ── Auto-forward delay (AWAITING_AF_DELAY) ────────────────────────────────
    if isinstance(state, str) and state == "AWAITING_AF_DELAY":
        user_states.pop(uid, None)
        try:
            secs = int(text.strip())
            if secs < 0: raise ValueError
            from database_dual import set_setting
            conn_id = context.user_data.get("af_editing_conn_id", "global")
            set_setting(f"af_delay_{conn_id}", str(secs))
            await safe_send_message(
                context.bot, chat_id,
                b(small_caps(f"✅ auto-forward delay set: {secs}s")),
                reply_markup=InlineKeyboardMarkup([[_back_btn("admin_autoforward"), _close_btn()]]),
            )
        except ValueError:
            await safe_send_message(context.bot, chat_id,
                b(small_caps("❌ invalid — send seconds (e.g. 5, 30, 60)")))
        return

    # ── Auto-forward caption override (AWAITING_AF_CAPTION) ──────────────────
    if isinstance(state, str) and state == "AWAITING_AF_CAPTION":
        user_states.pop(uid, None)
        cap = text.strip()
        from database_dual import set_setting
        conn_id = context.user_data.get("af_editing_conn_id", "global")
        set_setting(f"af_caption_{conn_id}", cap)
        label = e(cap[:60]) if cap else small_caps("cleared (using original)")
        await safe_send_message(
            context.bot, chat_id,
            b(small_caps("✅ auto-forward caption set:")) + "\n" + bq(label),
            reply_markup=InlineKeyboardMarkup([[_back_btn("admin_autoforward"), _close_btn()]]),
        )
        return

    # ── Auto-forward bulk message count (AWAITING_AF_BULK_COUNT) ─────────────
    if isinstance(state, str) and state == "AWAITING_AF_BULK_COUNT":
        user_states.pop(uid, None)
        try:
            count = int(text.strip())
            if count < 1: raise ValueError
            from database_dual import set_setting
            conn_id = context.user_data.get("af_editing_conn_id", "global")
            set_setting(f"af_bulk_count_{conn_id}", str(count))
            await safe_send_message(
                context.bot, chat_id,
                b(small_caps(f"✅ bulk forward count set: {count} messages")),
                reply_markup=InlineKeyboardMarkup([[_back_btn("admin_autoforward"), _close_btn()]]),
            )
        except ValueError:
            await safe_send_message(context.bot, chat_id,
                b(small_caps("❌ invalid — send a positive number")))
        return

    # ── Flood limit input (AWAITING_FLOOD_LIMIT) ──────────────────────────────
    if isinstance(state, str) and state == "AWAITING_FLOOD_LIMIT":
        user_states.pop(uid, None)
        try:
            limit = int(text.strip())
            if limit < 1: raise ValueError
            from database_dual import set_setting
            set_setting("flood_limit", str(limit))
            await safe_send_message(
                context.bot, chat_id,
                b(small_caps(f"✅ flood limit set: {limit} messages")),
                reply_markup=InlineKeyboardMarkup([[_back_btn("admin_spam_settings"), _close_btn()]]),
            )
        except ValueError:
            await safe_send_message(context.bot, chat_id,
                b(small_caps("❌ invalid — send a positive number (e.g. 5)")))
        return

    # ── Flood window input (AWAITING_FLOOD_WINDOW) ────────────────────────────
    if isinstance(state, str) and state == "AWAITING_FLOOD_WINDOW":
        user_states.pop(uid, None)
        try:
            secs = int(text.strip())
            if secs < 1: raise ValueError
            from database_dual import set_setting
            set_setting("flood_window_sec", str(secs))
            await safe_send_message(
                context.bot, chat_id,
                b(small_caps(f"✅ flood detection window set: {secs}s")),
                reply_markup=InlineKeyboardMarkup([[_back_btn("admin_spam_settings"), _close_btn()]]),
            )
        except ValueError:
            await safe_send_message(context.bot, chat_id,
                b(small_caps("❌ invalid — send seconds (e.g. 10)")))
        return

    # ── Flood ban duration (AWAITING_FLOOD_BAN_DUR) ───────────────────────────
    if isinstance(state, str) and state == "AWAITING_FLOOD_BAN_DUR":
        user_states.pop(uid, None)
        try:
            secs = int(text.strip())
            if secs < 0: raise ValueError
            from database_dual import set_setting
            set_setting("flood_ban_duration_sec", str(secs))
            label = f"{secs}s" if secs else "permanent"
            await safe_send_message(
                context.bot, chat_id,
                b(small_caps(f"✅ flood ban duration: {label}")),
                reply_markup=InlineKeyboardMarkup([[_back_btn("admin_rate_limit_settings"), _close_btn()]]),
            )
        except ValueError:
            await safe_send_message(context.bot, chat_id,
                b(small_caps("❌ invalid — send seconds (0 = permanent)")))
        return

    # ── Rate limit cooldown (AWAITING_RL_COOLDOWN) ────────────────────────────
    if isinstance(state, str) and state == "AWAITING_RL_COOLDOWN":
        user_states.pop(uid, None)
        try:
            secs = int(text.strip())
            if secs < 0: raise ValueError
            from database_dual import set_setting
            set_setting("rate_limit_cooldown_sec", str(secs))
            await safe_send_message(
                context.bot, chat_id,
                b(small_caps(f"✅ rate limit cooldown set: {secs}s")),
                reply_markup=InlineKeyboardMarkup([[_back_btn("admin_rate_limit_settings"), _close_btn()]]),
            )
        except ValueError:
            await safe_send_message(context.bot, chat_id,
                b(small_caps("❌ invalid — send seconds (e.g. 3)")))
        return

    # ── Search cache TTL (AWAITING_SEARCH_CACHE_TTL) ─────────────────────────
    if isinstance(state, str) and state == "AWAITING_SEARCH_CACHE_TTL":
        user_states.pop(uid, None)
        try:
            secs = int(text.strip())
            if secs < 0: raise ValueError
            from database_dual import set_setting
            set_setting("search_cache_ttl_sec", str(secs))
            label = f"{secs}s" if secs else "disabled"
            await safe_send_message(
                context.bot, chat_id,
                b(small_caps(f"✅ search cache TTL: {label}")),
                reply_markup=InlineKeyboardMarkup([[_back_btn("admin_search_settings"), _close_btn()]]),
            )
        except ValueError:
            await safe_send_message(context.bot, chat_id,
                b(small_caps("❌ invalid — send seconds (0 = no cache)")))
        return

    # ── Content blocklist words (AWAITING_BLOCKLIST_WORDS) ───────────────────
    if isinstance(state, str) and state == "AWAITING_BLOCKLIST_WORDS":
        user_states.pop(uid, None)
        val = text.strip()
        from database_dual import set_setting
        if val == "/clear":
            set_setting("content_blocklist_words", "")
            await safe_send_message(context.bot, chat_id,
                b(small_caps("✅ blocklist cleared")),
                reply_markup=InlineKeyboardMarkup([[_back_btn("admin_blocklist_settings"), _close_btn()]]))
        else:
            set_setting("content_blocklist_words", val)
            count = len([w.strip() for w in val.split(",") if w.strip()])
            await safe_send_message(
                context.bot, chat_id,
                b(small_caps(f"✅ {count} word(s) added to blocklist")),
                reply_markup=InlineKeyboardMarkup([[_back_btn("admin_blocklist_settings"), _close_btn()]]),
            )
        return

    # ── Default caption template (AWAITING_DEFAULT_CAPTION) ──────────────────
    if isinstance(state, str) and state == "AWAITING_DEFAULT_CAPTION":
        user_states.pop(uid, None)
        cap = text.strip()
        from database_dual import set_setting
        set_setting("default_caption_template", cap)
        await safe_send_message(
            context.bot, chat_id,
            b(small_caps("✅ default caption template saved")),
            reply_markup=InlineKeyboardMarkup([[_back_btn("admin_default_caption"), _close_btn()]]),
        )
        return

    # ── Welcome text (AWAITING_WELCOME_TEXT) ─────────────────────────────────
    if isinstance(state, str) and state == "AWAITING_WELCOME_TEXT":
        user_states.pop(uid, None)
        val = text.strip()
        if val:
            from database_dual import set_setting
            set_setting("welcome_text", val)
            await safe_send_message(
                context.bot, chat_id,
                b(small_caps("✅ welcome text saved")),
                reply_markup=InlineKeyboardMarkup([[_back_btn("admin_welcome_settings"), _close_btn()]]),
            )
        else:
            await safe_send_message(context.bot, chat_id, b(small_caps("❌ empty text not saved")))
        return

    # ── Welcome buttons (AWAITING_WELCOME_BUTTONS) ───────────────────────────
    if isinstance(state, str) and state == "AWAITING_WELCOME_BUTTONS":
        user_states.pop(uid, None)
        import json as _json_wb
        lines = [l.strip() for l in text.strip().splitlines() if l.strip()]
        btns = []
        errors = []
        for ln in lines:
            if " - " in ln:
                parts = ln.split(" - ", 1)
                btns.append({"text": parts[0].strip(), "url": parts[1].strip()})
            else:
                errors.append(ln)
        from database_dual import set_setting
        set_setting("welcome_buttons", _json_wb.dumps(btns))
        msg = b(small_caps(f"✅ {len(btns)} button(s) saved"))
        if errors:
            msg += "\n" + bq(small_caps(f"skipped {len(errors)} invalid line(s)"))
        await safe_send_message(context.bot, chat_id, msg,
            reply_markup=InlineKeyboardMarkup([[_back_btn("admin_welcome_settings"), _close_btn()]]))
        return

    # ── Welcome media (AWAITING_WELCOME_MEDIA) ────────────────────────────────
    if isinstance(state, str) and state == "AWAITING_WELCOME_MEDIA":
        user_states.pop(uid, None)
        from database_dual import set_setting
        if message and message.photo:
            fid = message.photo[-1].file_id
            set_setting("welcome_media_url", fid)
            await safe_send_message(context.bot, chat_id,
                b(small_caps("✅ welcome photo saved")),
                reply_markup=InlineKeyboardMarkup([[_back_btn("admin_welcome_settings"), _close_btn()]]))
        elif message and message.animation:
            fid = message.animation.file_id
            set_setting("welcome_media_url", fid)
            await safe_send_message(context.bot, chat_id,
                b(small_caps("✅ welcome gif/animation saved")),
                reply_markup=InlineKeyboardMarkup([[_back_btn("admin_welcome_settings"), _close_btn()]]))
        elif message and message.video:
            fid = message.video.file_id
            set_setting("welcome_media_url", fid)
            await safe_send_message(context.bot, chat_id,
                b(small_caps("✅ welcome video saved")),
                reply_markup=InlineKeyboardMarkup([[_back_btn("admin_welcome_settings"), _close_btn()]]))
        elif text and text.strip() == "/clear":
            set_setting("welcome_media_url", "")
            await safe_send_message(context.bot, chat_id,
                b(small_caps("✅ welcome media cleared")),
                reply_markup=InlineKeyboardMarkup([[_back_btn("admin_welcome_settings"), _close_btn()]]))
        else:
            await safe_send_message(context.bot, chat_id,
                b(small_caps("❌ send a photo, gif, video, or /clear to remove")))
        return

    # ── Goodbye text (AWAITING_GOODBYE_TEXT) ─────────────────────────────────
    if isinstance(state, str) and state == "AWAITING_GOODBYE_TEXT":
        user_states.pop(uid, None)
        val = text.strip()
        if val:
            from database_dual import set_setting
            set_setting("goodbye_text", val)
            await safe_send_message(
                context.bot, chat_id,
                b(small_caps("✅ goodbye text saved")),
                reply_markup=InlineKeyboardMarkup([[_back_btn("admin_goodbye_settings"), _close_btn()]]),
            )
        else:
            await safe_send_message(context.bot, chat_id, b(small_caps("❌ empty text not saved")))
        return

    # ── Maintenance broadcast message (AWAITING_MAINTENANCE_MESSAGE) ─────────
    if isinstance(state, str) and state == "AWAITING_MAINTENANCE_MESSAGE":
        user_states.pop(uid, None)
        msg_text = text.strip()
        if not msg_text:
            await safe_send_message(context.bot, chat_id, b(small_caps("❌ empty message not sent")))
            return
        from database_dual import set_setting
        set_setting("maintenance_message", msg_text)
        await safe_send_message(
            context.bot, chat_id,
            b(small_caps("✅ maintenance message saved and will show to users during maintenance")),
            reply_markup=InlineKeyboardMarkup([[_back_btn("admin_settings"), _close_btn()]]),
        )
        return

    # ── Poster font primary (AWAITING_POSTER_FONT_PRIMARY) ───────────────────
    if isinstance(state, str) and state == "AWAITING_POSTER_FONT_PRIMARY":
        user_states.pop(uid, None)
        val = text.strip()
        if val:
            from database_dual import set_setting
            set_setting("poster_font_primary", val)
            await safe_send_message(context.bot, chat_id,
                b(small_caps(f"✅ title font set: {e(val)}")),
                reply_markup=InlineKeyboardMarkup([[_back_btn("admin_poster_text_settings"), _close_btn()]]))
        else:
            await safe_send_message(context.bot, chat_id, b(small_caps("❌ empty — not saved")))
        return

    # ── Poster font secondary (AWAITING_POSTER_FONT_SECONDARY) ───────────────
    if isinstance(state, str) and state == "AWAITING_POSTER_FONT_SECONDARY":
        user_states.pop(uid, None)
        val = text.strip()
        if val:
            from database_dual import set_setting
            set_setting("poster_font_secondary", val)
            await safe_send_message(context.bot, chat_id,
                b(small_caps(f"✅ info font set: {e(val)}")),
                reply_markup=InlineKeyboardMarkup([[_back_btn("admin_poster_text_settings"), _close_btn()]]))
        else:
            await safe_send_message(context.bot, chat_id, b(small_caps("❌ empty — not saved")))
        return

    # ── Poster title color (AWAITING_POSTER_TITLE_COLOR) ─────────────────────
    if isinstance(state, str) and state == "AWAITING_POSTER_TITLE_COLOR":
        user_states.pop(uid, None)
        import re as _re_color
        val = text.strip()
        if _re_color.match(r"^#[0-9A-Fa-f]{6}$", val):
            from database_dual import set_setting
            set_setting("poster_color_title", val)
            await safe_send_message(context.bot, chat_id,
                b(small_caps(f"✅ title color set: {e(val)}")),
                reply_markup=InlineKeyboardMarkup([[_back_btn("admin_poster_text_settings"), _close_btn()]]))
        else:
            await safe_send_message(context.bot, chat_id,
                b(small_caps("❌ invalid hex color — send like #FFFFFF or #FFD700")))
        return

    # ── Prune threshold days (AWAITING_PRUNE_THRESHOLD) ──────────────────────
    if isinstance(state, str) and state == "AWAITING_PRUNE_THRESHOLD":
        user_states.pop(uid, None)
        try:
            days = int(text.strip())
            if days < 1: raise ValueError
            from database_dual import set_setting
            set_setting("prune_inactive_days", str(days))
            await safe_send_message(
                context.bot, chat_id,
                b(small_caps(f"✅ prune threshold: {days} days")),
                reply_markup=InlineKeyboardMarkup([[_back_btn("admin_prune_users"), _close_btn()]]),
            )
        except ValueError:
            await safe_send_message(context.bot, chat_id,
                b(small_caps("❌ invalid — send a positive number of days")))
        return


    logger.debug(f"Admin message in unknown state {state} from {uid}: {text[:50]}")
